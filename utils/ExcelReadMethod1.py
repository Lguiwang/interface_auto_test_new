# -*- coding:utf-8 -*-
"""
======================
    Time  ：2024/7/24 15:21
    Author: Liuguiwang
    File  : ExcelReadMethod1.py  直接获取到每条用例，每条用例放入列表中，再放入总列表中
======================
"""
import openpyxl
import os

from utils.LogUtil import MyLog

log = MyLog().getLog()


class ExcelReader(object):
    # 判断要读取的文件是否存在
    def __init__(self, excel_name, sheet_name):
        try:
            self.excel_name = excel_name
            self.sheet_name = sheet_name
        except FileNotFoundError as e:
            log.error(e)

    # 获取全部测试用例
    def get_all_cases(self):
        # 打开表格
        workbook = openpyxl.load_workbook(self.excel_name)
        # 选择表格文件，哪一个sheet表
        sheet = workbook[self.sheet_name]
        # 创建获取到用例存放的列表
        cases_list = list()
        # 若有标题+表头，则从第3行开始获取，只有表头，从第2行获取, openpyxl遍历表格时下标从1开始
        for row in range(2, sheet.max_row + 1):
            line = list()
            # 获取每一行用例信息存放到列表中
            for column in range(1, sheet.max_column + 1):
                line.append(sheet.cell(row, column).value)
            cases_list.append(line)
        return cases_list

    # 遍历总用例列表，查找前置用例，pre为前置用例的用例ID（能唯一区分用例的字段即可）
    def get_pre_case(self, pre):
        cases_list = self.get_all_cases()
        for case in cases_list:
            if pre in case:
                # 返回符合条件的前置用例
                return case
        else:
            return None


if __name__ == '__main__':
    print(ExcelReader(os.path.dirname(os.getcwd()) + '//data//test_excel.xlsx',
                      'Sheet1').get_all_cases())